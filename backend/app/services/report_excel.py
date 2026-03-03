"""
报告 Excel 构建模块
使用 openpyxl 生成多 Sheet Excel 并嵌入图表图片
"""
from io import BytesIO
from datetime import datetime
from typing import List, Dict, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage

# ---------------------------------------------------------------------------
# 样式常量
# ---------------------------------------------------------------------------
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
TITLE_FONT = Font(bold=True, size=16)
SUBTITLE_FONT = Font(bold=True, size=12)
META_LABEL_FONT = Font(bold=True, size=11)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)
NUMBER_FMT = '#,##0'
PCT_FMT = '0.0%'
DPPM_FMT = '#,##0'
INVALID_FILENAME_CHARS = '\\/:*?"<>|'


class ReportExcelBuilder:
    """报告 Excel 构建器"""

    def __init__(self, kpi_type: str, dimension: str):
        self.kpi_type = kpi_type
        self.dimension = dimension
        self.wb = Workbook()
        self.wb.remove(self.wb.active)

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------
    def _write_header(self, ws, row: int, headers: List[str]):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER

    def _write_row(self, ws, row: int, values: list, fmt_map: Optional[Dict[int, str]] = None):
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=v)
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            if fmt_map and col in fmt_map:
                cell.number_format = fmt_map[col]

    def _auto_width(self, ws, min_width: int = 12, max_width: int = 40):
        for col_cells in ws.columns:
            max_len = min_width
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)) + 4, max_width))
            ws.column_dimensions[col_letter].width = max_len

    def _embed_image(self, ws, png_bytes: bytes, anchor: str,
                     width: int = 800, height: int = 400):
        img = XlImage(BytesIO(png_bytes))
        img.width = width
        img.height = height
        ws.add_image(img, anchor)

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------
    def add_info_sheet(self, meta: dict, tgt: Optional[int] = None):
        ws = self.wb.create_sheet("报告信息")
        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 60

        title = f"{self.kpi_type} {self.dimension}分析报告"
        cell = ws.cell(row=1, column=1, value=title)
        cell.font = TITLE_FONT

        rows = [
            ("数据截至", meta.get("data_as_of", "")),
            ("分析时间范围", meta.get("time_range", "")),
            (f"分析{self.dimension}", meta.get("entities", "")),
        ]
        extra_filters = meta.get("extra_filters") or {}
        for label, val in extra_filters.items():
            rows.append((label, val))
        if tgt is not None:
            rows.append(("TGT目标", f"{tgt:,} DPPM"))
        rows.append(("报告生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")))

        for i, (label, val) in enumerate(rows, 3):
            ws.cell(row=i, column=1, value=label).font = META_LABEL_FONT
            ws.cell(row=i, column=2, value=val)

    def add_trend_sheet(
        self,
        entities: List[Dict],
        chart_png: Optional[bytes] = None,
        value_key: str = "ifir",
        unit_label: str = "IFIR",
    ):
        ws = self.wb.create_sheet("趋势数据")
        all_months = sorted({p["month"] for e in entities for p in e["trend"]})
        headers = ["Month"] + [f'{e["name"]} {unit_label}(DPPM)' for e in entities]
        self._write_header(ws, 1, headers)

        for r_idx, month in enumerate(all_months, 2):
            ws.cell(row=r_idx, column=1, value=month).border = THIN_BORDER
            for c_idx, entity in enumerate(entities, 2):
                month_map = {p["month"]: p["value"] for p in entity["trend"]}
                raw = month_map.get(month)
                dppm = round(raw * 1_000_000) if raw is not None else None
                cell = ws.cell(row=r_idx, column=c_idx, value=dppm)
                cell.border = THIN_BORDER
                cell.number_format = DPPM_FMT

        self._auto_width(ws)
        if chart_png:
            anchor_row = len(all_months) + 4
            self._embed_image(ws, chart_png, f"A{anchor_row}")

    def add_top_issue_sheet(self, cards_data: List[Dict], entity_key: str = "model"):
        ws = self.wb.create_sheet("Top Issue汇总")
        headers = [self.dimension, "Rank", "Issue", "Count", "Share(%)"]
        row = 1
        self._write_header(ws, row, headers)
        row += 1

        for card in cards_data:
            name = card.get(entity_key, "")
            for issue in card.get("top_issues", []):
                self._write_row(ws, row, [
                    name, issue.get("rank"), issue.get("issue"),
                    issue.get("count"),
                    issue.get("share"),
                ], fmt_map={5: PCT_FMT})
                row += 1
        self._auto_width(ws)

    def add_monthly_top_issue_sheet(self, cards_data: List[Dict], entity_key: str = "model"):
        ws = self.wb.create_sheet("月度Top Issue")
        headers = ["Month", self.dimension, "Rank", "Issue", "Count", "Share(%)"]
        row = 1
        self._write_header(ws, row, headers)
        row += 1

        for card in cards_data:
            name = card.get(entity_key, "")
            for monthly in card.get("monthly_top_issues", []) or []:
                month = monthly.get("month", "")
                for issue in monthly.get("items", []):
                    self._write_row(ws, row, [
                        month, name, issue.get("rank"), issue.get("issue"),
                        issue.get("count"), issue.get("share"),
                    ], fmt_map={6: PCT_FMT})
                    row += 1
        self._auto_width(ws)

    def add_top_model_sheet(self, cards_data: List[Dict], entity_key: str = "odm",
                            value_label: str = "IFIR",
                            claim_key: str = "box_claim", mm_key: str = "box_mm"):
        ws = self.wb.create_sheet("Top Model汇总")
        headers = [self.dimension, "Rank", "Model", "Top Issues",
                   f"{value_label}(DPPM)", claim_key.upper(), mm_key.upper()]
        row = 1
        self._write_header(ws, row, headers)
        row += 1

        for card in cards_data:
            name = card.get(entity_key, "")
            for m in card.get("top_models", []):
                issues_str = self._format_top_issues(m.get("top_issues"))
                kpi_val = m.get(value_label.lower()) or m.get("ifir") or m.get("ra") or 0
                dppm = round(kpi_val * 1_000_000)
                self._write_row(ws, row, [
                    name, m.get("rank"), m.get("model"), issues_str,
                    dppm, m.get(claim_key), m.get(mm_key),
                ], fmt_map={5: DPPM_FMT, 6: NUMBER_FMT, 7: NUMBER_FMT})
                row += 1
        self._auto_width(ws)

    def add_monthly_top_model_sheet(self, cards_data: List[Dict], entity_key: str = "odm",
                                     value_label: str = "IFIR",
                                     claim_key: str = "box_claim", mm_key: str = "box_mm"):
        ws = self.wb.create_sheet("月度Top Model")
        headers = ["Month", self.dimension, "Rank", "Model", "Top Issues",
                   f"{value_label}(DPPM)", claim_key.upper(), mm_key.upper()]
        row = 1
        self._write_header(ws, row, headers)
        row += 1

        for card in cards_data:
            name = card.get(entity_key, "")
            for monthly in card.get("monthly_top_models", []) or []:
                month = monthly.get("month", "")
                for m in monthly.get("items", []):
                    issues_str = self._format_top_issues(m.get("top_issues"))
                    kpi_val = m.get(value_label.lower()) or m.get("ifir") or m.get("ra") or 0
                    dppm = round(kpi_val * 1_000_000)
                    self._write_row(ws, row, [
                        month, name, m.get("rank"), m.get("model"), issues_str,
                        dppm, m.get(claim_key), m.get(mm_key),
                    ], fmt_map={6: DPPM_FMT, 7: NUMBER_FMT, 8: NUMBER_FMT})
                    row += 1
        self._auto_width(ws)

    def add_top_odm_sheet(self, cards_data: List[Dict],
                          value_label: str = "IFIR",
                          claim_key: str = "box_claim", mm_key: str = "box_mm"):
        ws = self.wb.create_sheet("Top ODM汇总")
        headers = ["Segment", "Rank", "ODM",
                   f"{value_label}(DPPM)", claim_key.upper(), mm_key.upper()]
        row = 1
        self._write_header(ws, row, headers)
        row += 1

        for card in cards_data:
            seg = card.get("segment", "")
            for o in card.get("top_odms", []):
                kpi_val = o.get(value_label.lower()) or o.get("ifir") or o.get("ra") or 0
                dppm = round(kpi_val * 1_000_000)
                self._write_row(ws, row, [
                    seg, o.get("rank"), o.get("odm"),
                    dppm, o.get(claim_key), o.get(mm_key),
                ], fmt_map={4: DPPM_FMT, 5: NUMBER_FMT, 6: NUMBER_FMT})
                row += 1
        self._auto_width(ws)

    def add_monthly_top_odm_sheet(self, cards_data: List[Dict],
                                   value_label: str = "IFIR",
                                   claim_key: str = "box_claim", mm_key: str = "box_mm"):
        ws = self.wb.create_sheet("月度Top ODM")
        headers = ["Month", "Segment", "Rank", "ODM",
                   f"{value_label}(DPPM)", claim_key.upper(), mm_key.upper()]
        row = 1
        self._write_header(ws, row, headers)
        row += 1

        for card in cards_data:
            seg = card.get("segment", "")
            for monthly in card.get("monthly_top_odms", []) or []:
                month = monthly.get("month", "")
                for o in monthly.get("items", []):
                    kpi_val = o.get(value_label.lower()) or o.get("ifir") or o.get("ra") or 0
                    dppm = round(kpi_val * 1_000_000)
                    self._write_row(ws, row, [
                        month, seg, o.get("rank"), o.get("odm"),
                        dppm, o.get(claim_key), o.get(mm_key),
                    ], fmt_map={5: DPPM_FMT, 6: NUMBER_FMT, 7: NUMBER_FMT})
                    row += 1
        self._auto_width(ws)

    def add_detail_sheet(
        self,
        detail_rows: List[Dict],
        columns: List[Tuple[str, str]],
        total: int = 0,
        truncated: bool = False,
    ):
        ws = self.wb.create_sheet("Detail明细数据")
        headers = [display for display, _ in columns]
        self._write_header(ws, 1, headers)

        for r_idx, row_data in enumerate(detail_rows, 2):
            for c_idx, (_, db_key) in enumerate(columns, 1):
                val = row_data.get(db_key)
                if isinstance(val, (bytes, bytearray)):
                    val = val.decode("utf-8", errors="replace")
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = THIN_BORDER

        self._auto_width(ws)

        if truncated:
            note_row = len(detail_rows) + 3
            cell = ws.cell(row=note_row, column=1,
                           value=f"数据量过大，仅导出前 {len(detail_rows):,} 行（共约 {total:,} 行），完整数据请联系管理员。")
            cell.font = Font(color="FF0000", italic=True)

    def add_comparison_sheet(
        self,
        pie_data: List[Dict],
        chart_png: Optional[bytes] = None,
        entity_key: str = "name",
        value_label: str = "IFIR",
        claim_key: str = "box_claim",
        mm_key: str = "box_mm",
    ):
        ws = self.wb.create_sheet(f"多{self.dimension}对比")
        headers = [self.dimension, f"{value_label}(DPPM)", "Share(%)",
                   claim_key.upper(), mm_key.upper()]
        self._write_header(ws, 1, headers)

        for r_idx, item in enumerate(pie_data, 2):
            self._write_row(ws, r_idx, [
                item.get(entity_key, ""),
                item.get("dppm") or round((item.get(value_label.lower(), 0) or 0) * 1_000_000),
                item.get("share"),
                item.get(claim_key),
                item.get(mm_key),
            ], fmt_map={2: DPPM_FMT, 3: PCT_FMT, 4: NUMBER_FMT, 5: NUMBER_FMT})

        self._auto_width(ws)
        if chart_png:
            anchor_row = len(pie_data) + 4
            self._embed_image(ws, chart_png, f"A{anchor_row}", width=640, height=480)

    # ------------------------------------------------------------------
    # util
    # ------------------------------------------------------------------
    @staticmethod
    def _format_top_issues(issues) -> str:
        if not issues:
            return ""
        parts = []
        for iss in issues:
            if isinstance(iss, dict):
                parts.append(f"{iss.get('issue', '')}*{iss.get('count', '')}")
            else:
                parts.append(f"{iss.issue}*{iss.count}")
        return ", ".join(parts)

    # ------------------------------------------------------------------
    # save
    # ------------------------------------------------------------------
    def save_to_file(self, file_path: str) -> str:
        self.wb.save(file_path)
        return file_path


# ---------------------------------------------------------------------------
# 文件名构建
# ---------------------------------------------------------------------------
def build_report_filename(
    kpi_type: str,
    dimension: str,
    entities: List[str],
    start_month: str,
    end_month: str,
) -> str:
    def sanitize_part(value: str) -> str:
        cleaned = "".join("_" if ch in INVALID_FILENAME_CHARS else ch for ch in value)
        return cleaned.replace(" ", "_")

    if len(entities) <= 3:
        name_part = "+".join(sanitize_part(e) for e in entities)
    else:
        name_part = "+".join(sanitize_part(e) for e in entities[:2]) + f"等{len(entities)}个"
    date_str = datetime.now().strftime("%Y%m%d")
    return f"{kpi_type}_{dimension}分析报告_{name_part}_{start_month}至{end_month}_{date_str}.xlsx"
